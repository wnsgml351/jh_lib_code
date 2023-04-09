import requests
from bs4 import BeautifulSoup
import pyautogui
from openpyxl import Workbook
from openpyxl.styles import Alignment

# 키워드 입력
keyword = pyautogui.prompt("검색어를 입력하세요.")

# 엑셀 파일 객체 생성
wb = Workbook()

# 이름이 있는 시트를 생성
ws = wb.create_sheet(keyword)

# 칼럼 길이 조정
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 120

# 행 번호
row = 1

response = requests.get(f"https://search.naver.com/search.naver?where=news&sm=tab_jum&query={keyword}")
html = response.text
soup = BeautifulSoup(html, 'html.parser')
articles = soup.select("div.info_group") # 뉴스 기사 div 10개 추출
for article in articles:
    links = article.select("a.info")
    if len(links) >= 2: # 링크가 2개 이상이면
        url = links[1].attrs['href'] # 두번째 링크의 href 추출
        # 다시 request를 날려 준다
        response = requests.get(url, headers={'User-Agent' : 'Mozila/5.0'})
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')
        
        # 연예 뉴스 체크
        if "entertain" in response.url: 
            title = soup.select_one(".end_tit")
            content = soup.select_one("#articeBody")
        elif "sports" in response.url:
            title = soup.select_one("h4.title")
            content = soup.select_one("#newsEndContents")

            # 본문 내용 중에 불 필요한 div 제거
            divs = content.select("div")
            for div in divs:
                div.decompose()
            
            # 본문 내용 중에 불 필요한 p 제거
            paragraphs = content.select("p")
            for p in paragraphs:
                p.decompose()

        else: 
            title = soup.select_one(".media_end_head_headline")
            content = soup.select_one("#newsct_article")

        print("=======링크======= \n", url)
        print("=======제목======= \n", title.text.strip())
        print("=======본문======= \n", content.text.strip())

        ws[f'A{row}'] = url
        ws[f'B{row}'] = title.text.strip()
        ws[f'C{row}'] = content.text.strip()

        # 자동 줄 바꿈
        ws[f'C{row}'].alignment = Alignment(wrap_text=True)
        
        
        row = row + 1

wb.save(f'{keyword}_result.xlsx')
